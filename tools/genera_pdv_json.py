from pathlib import Path
import json
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
SRC = ROOT / 'Dati' / 'PDV-e-ESATTORI.xlsx'
DST = ROOT / 'Dati' / 'PDV-e-ESATTORI.json'

OVERRIDES = {
    'BAR MATCH SRLS': 'ALESSANDRO',
    'MALIBU BAR DI AYALA ARANA ELSY MARIANA': 'MATTEO',
    'MAX BAR': 'SANREMO',
    'TABACCHERIA BADALUCCO': 'ALESSANDRO',
}

def clean(v):
    return ' '.join(str(v or '').strip().split())

wb = load_workbook(SRC, read_only=True, data_only=True)
ws = max(wb.worksheets, key=lambda s: s.max_row)
rows = ws.iter_rows(values_only=True)
headers = [clean(v) for v in next(rows)]
try:
    sede_i = headers.index('DENOMIN. SEDE')
    esa_i = headers.index('ESATTORE')
except ValueError as exc:
    raise SystemExit(f'Colonne richieste non trovate. Intestazioni: {headers}') from exc

mapping = {}
for row in rows:
    sede = clean(row[sede_i] if sede_i < len(row) else '')
    esa = clean(row[esa_i] if esa_i < len(row) else '').upper()
    if not sede:
        continue
    if esa == 'LUCA':
        esa = 'SANREMO'
    mapping[sede] = esa or '—'

for sede, esa in OVERRIDES.items():
    mapping[sede] = esa

out = [
    {'DENOMIN. SEDE': sede, 'ESATTORE': mapping[sede]}
    for sede in sorted(mapping, key=lambda x: x.upper())
]
DST.write_text(json.dumps(out, ensure_ascii=False, indent=2) + '\n', encoding='utf-8')
print(f'Scritte {len(out)} associazioni in {DST.relative_to(ROOT)}')
